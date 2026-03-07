from __future__ import annotations

import csv
import datetime as dt
import io
import os
import sqlite3
import tempfile
import requests
import re
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup


from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from openpyxl import load_workbook
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# =========================================================
# OMPP / Observatorio Inteligente de Precios y Abastecimiento
# main.py - versión integral para pegar y correr
# ---------------------------------------------------------
# Qué incluye:
# - Dashboard institucional
# - Importación segura de Excel hoja Canasta_25
# - Histórico semanal en SQLite
# - Ranking semanal
# - Reporte PDF
# - Backup CSV
# - API de resumen y ranking
# - Módulo IPPS (Índice de Presión de Precios Semanal)
# - Módulo de alertas tempranas
# - Módulo territorial (base lista para crecer)
# - Módulo social (presión sobre ingreso)
# - Indicadores internacionales actualizables por API
# - Migraciones automáticas básicas de esquema
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "ompp.db"
UPLOAD_DIR = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

APP_TITLE = "OMPP Sistema con Reporte Real — Dashboard"
DEFAULT_SOURCE = "Excel"
JOB_TOKEN = os.getenv("JOB_TOKEN", "")

app = FastAPI(title=APP_TITLE)


# =========================================================
# CONFIGURACIÓN BASE
# =========================================================
DEFAULT_PRODUCTS = [
    ("Arroz", 5.0, 0.60, 0.30, 0.20),
    ("Aceite", 4.0, 0.75, 0.45, 0.30),
    ("Azúcar", 3.0, 0.35, 0.20, 0.20),
    ("Yerba", 3.0, 0.20, 0.15, 0.15),
    ("Fideo", 3.0, 0.70, 0.30, 0.20),
    ("Harina", 4.0, 0.80, 0.25, 0.15),
    ("Pan", 5.0, 0.55, 0.30, 0.20),
    ("Leche", 5.0, 0.35, 0.25, 0.20),
    ("Huevos", 4.0, 0.10, 0.25, 0.20),
    ("Queso", 4.0, 0.20, 0.25, 0.20),
    ("Pollo", 5.0, 0.15, 0.30, 0.25),
    ("Carne vacuna", 7.0, 0.05, 0.35, 0.30),
    ("Cerdo", 3.0, 0.10, 0.30, 0.25),
    ("Tomate", 5.0, 0.10, 0.35, 0.40),
    ("Cebolla", 4.0, 0.35, 0.30, 0.35),
    ("Papa", 4.0, 0.40, 0.25, 0.30),
    ("Locote", 2.0, 0.15, 0.30, 0.35),
    ("Banana", 2.0, 0.05, 0.15, 0.10),
    ("Naranja", 2.0, 0.05, 0.15, 0.10),
    ("Mandioca", 2.0, 0.01, 0.20, 0.20),
    ("Detergente", 2.0, 0.50, 0.20, 0.15),
    ("Jabón", 2.0, 0.45, 0.20, 0.15),
    ("Papel higiénico", 2.0, 0.35, 0.15, 0.10),
    ("Gas doméstico", 4.0, 0.85, 0.70, 0.80),
    ("Agua mineral", 2.0, 0.05, 0.15, 0.15),
]

DEFAULT_LOCATIONS = [
    ("NAT", "Nacional", "Nacional", "promedio", 1),
    ("ASU", "Asunción", "Asunción", "urbano", 1),
    ("CEN", "Central", "Central", "urbano", 1),
    ("FRO", "Frontera", "Frontera", "frontera", 1),
    ("INT", "Interior", "Interior", "regional", 1),
]

DEFAULT_INDICATORS = [
    ("usd_pyg", "USD / PYG"),
    ("brent", "Brent"),
    ("diesel", "Diesel"),
    ("gasolina", "Gasolina"),
    ("trigo", "Trigo"),
    ("maiz", "Maíz"),
]


@contextmanager
def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# =========================================================
# HELPERS
# =========================================================
def normalize_name(value: str | None) -> str:
    return " ".join((value or "").strip().split()).lower()



def iso_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return dt.datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                pass
    return None



def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip().replace(" ", "")
        if not txt:
            return None
        if txt.count(",") == 1 and txt.count(".") >= 1:
            txt = txt.replace(".", "").replace(",", ".")
        elif txt.count(",") == 1 and txt.count(".") == 0:
            txt = txt.replace(",", ".")
        else:
            txt = txt.replace(",", "")
        try:
            return float(txt)
        except ValueError:
            return None
    return None



def money(v: float | None) -> str:
    if v is None:
        return "—"
    return f"Gs. {v:,.0f}".replace(",", ".")



def pct(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{v:+.2f}%"



def safe_div(n: float | None, d: float | None) -> float | None:
    if n is None or d in (None, 0):
        return None
    return n / d



def calculate_variation(current: float | None, previous: float | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return ((current - previous) / previous) * 100.0


# =========================================================
# BASE DE DATOS / MIGRACIONES
# =========================================================
def table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {r[1] for r in rows}



def ensure_column(conn: sqlite3.Connection, table_name: str, definition: str, column_name: str) -> None:
    cols = table_columns(conn, table_name)
    if column_name not in cols:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {definition}")



def init_db() -> None:
    with db() as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                name TEXT NOT NULL,
                normalized_name TEXT NOT NULL UNIQUE,
                unit TEXT DEFAULT 'unidad',
                weight REAL DEFAULT 1,
                import_dependency REAL DEFAULT 0,
                fx_sensitivity REAL DEFAULT 0,
                fuel_sensitivity REAL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS locations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                region_type TEXT DEFAULT 'general',
                market_type TEXT DEFAULT 'promedio',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS weekly_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_date TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                location_id INTEGER,
                price REAL NOT NULL,
                source TEXT NOT NULL DEFAULT 'Excel',
                imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(week_date, product_id, location_id),
                FOREIGN KEY(product_id) REFERENCES products(id),
                FOREIGN KEY(location_id) REFERENCES locations(id)
            );

            CREATE TABLE IF NOT EXISTS indicators (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                key TEXT NOT NULL UNIQUE,
                label TEXT NOT NULL,
                value REAL,
                variation REAL,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS income_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                label TEXT NOT NULL DEFAULT 'Salario mínimo mensual',
                monthly_income REAL NOT NULL DEFAULT 2798510,
                household_size REAL DEFAULT 4,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS imports_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                status TEXT NOT NULL,
                message TEXT,
                rows_detected INTEGER DEFAULT 0,
                rows_inserted INTEGER DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            """
        )

        # migraciones simples para bases viejas
        ensure_column(conn, "products", "weight REAL DEFAULT 1", "weight")
        ensure_column(conn, "products", "import_dependency REAL DEFAULT 0", "import_dependency")
        ensure_column(conn, "products", "fx_sensitivity REAL DEFAULT 0", "fx_sensitivity")
        ensure_column(conn, "products", "fuel_sensitivity REAL DEFAULT 0", "fuel_sensitivity")

        wp_cols = table_columns(conn, "weekly_prices")
        if "location_id" not in wp_cols:
            conn.execute("ALTER TABLE weekly_prices ADD COLUMN location_id INTEGER")

        # seed products
        cur = conn.execute("SELECT COUNT(*) AS n FROM products")
        if cur.fetchone()["n"] == 0:
            for idx, item in enumerate(DEFAULT_PRODUCTS, start=1):
                name, weight, import_dep, fx_sens, fuel_sens = item
                conn.execute(
                    """
                    INSERT INTO products (
                        code, name, normalized_name, unit, weight,
                        import_dependency, fx_sensitivity, fuel_sensitivity, active
                    )
                    VALUES (?, ?, ?, 'unidad', ?, ?, ?, ?, 1)
                    """,
                    (
                        f"P{idx:02d}",
                        name,
                        normalize_name(name),
                        weight,
                        import_dep,
                        fx_sens,
                        fuel_sens,
                    ),
                )
        else:
            # completa metadata faltante si existía la tabla anterior
            for item in DEFAULT_PRODUCTS:
                name, weight, import_dep, fx_sens, fuel_sens = item
                conn.execute(
                    """
                    UPDATE products
                    SET weight = COALESCE(weight, ?),
                        import_dependency = COALESCE(import_dependency, ?),
                        fx_sensitivity = COALESCE(fx_sensitivity, ?),
                        fuel_sensitivity = COALESCE(fuel_sensitivity, ?)
                    WHERE normalized_name = ?
                    """,
                    (weight, import_dep, fx_sens, fuel_sens, normalize_name(name)),
                )

        # seed locations
        for code, name, region_type, market_type, active in DEFAULT_LOCATIONS:
            conn.execute(
                """
                INSERT INTO locations (code, name, region_type, market_type, active)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(code) DO NOTHING
                """,
                (code, name, region_type, market_type, active),
            )

        # si hay registros viejos sin location_id, los pasa a NAT
        nat = conn.execute("SELECT id FROM locations WHERE code='NAT'").fetchone()
        if nat:
            conn.execute(
                "UPDATE weekly_prices SET location_id = ? WHERE location_id IS NULL",
                (int(nat["id"]),),
            )

        # seed indicators
        for key, label in DEFAULT_INDICATORS:
            conn.execute(
                "INSERT INTO indicators (key, label) VALUES (?, ?) ON CONFLICT(key) DO NOTHING",
                (key, label),
            )

        # seed income settings
        cur = conn.execute("SELECT COUNT(*) AS n FROM income_settings")
        if cur.fetchone()["n"] == 0:
            conn.execute(
                "INSERT INTO income_settings (label, monthly_income, household_size) VALUES (?, ?, ?)",
                ("Salario mínimo mensual", 2798510, 4),
            )


@app.on_event("startup")
def on_startup() -> None:
    init_db()


# =========================================================
# REPOSITORIO
# =========================================================
def fetch_products(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT id, code, name, unit, normalized_name, weight,
               import_dependency, fx_sensitivity, fuel_sensitivity
        FROM products
        WHERE active=1
        ORDER BY id
        """
    ).fetchall()



def fetch_locations(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT id, code, name, region_type, market_type FROM locations WHERE active=1 ORDER BY id"
    ).fetchall()



def get_location_id(conn: sqlite3.Connection, code: str = "NAT") -> int:
    row = conn.execute("SELECT id FROM locations WHERE code=?", (code,)).fetchone()
    if not row:
        raise ValueError(f"No existe la ubicación {code}")
    return int(row["id"])



def latest_week(conn: sqlite3.Connection, location_code: str = "NAT") -> str | None:
    row = conn.execute(
        """
        SELECT MAX(w.week_date) AS d
        FROM weekly_prices w
        JOIN locations l ON l.id = w.location_id
        WHERE l.code = ?
        """,
        (location_code,),
    ).fetchone()
    return row["d"] if row and row["d"] else None



def previous_week(conn: sqlite3.Connection, week_date: str | None, location_code: str = "NAT") -> str | None:
    if not week_date:
        return None
    row = conn.execute(
        """
        SELECT MAX(w.week_date) AS d
        FROM weekly_prices w
        JOIN locations l ON l.id = w.location_id
        WHERE l.code = ? AND w.week_date < ?
        """,
        (location_code, week_date),
    ).fetchone()
    return row["d"] if row and row["d"] else None



def indicators_map(conn: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        "SELECT key, label, value, variation, updated_at FROM indicators ORDER BY id"
    ).fetchall()
    return {
        r["key"]: {
            "label": r["label"],
            "value": r["value"],
            "variation": r["variation"],
            "updated_at": r["updated_at"],
        }
        for r in rows
    }
def fetch_usd_pyg_bcp() -> dict[str, float | str | None]:
    """
    Lee el tipo de cambio referencial USD/PYG desde la página del BCP.
    """

    url = "https://www.bcp.gov.py/webapps/web/cotizacion/monedas"

    try:
        resp = requests.get(url, timeout=20)
        resp.raise_for_status()
        html = resp.text

        # Patrón más simple y más robusto para la línea del dólar
        match = re.search(
            r"D[ÓO]LAR\s+ESTADOUNIDENSE\s+USD\s+1,0000\s+([0-9.,]+)",
            html,
            re.IGNORECASE,
        )

        if not match:
            return {"value": None, "variation": None, "source": "BCP"}

        raw_value = match.group(1).strip()
        value = float(raw_value.replace(".", "").replace(",", "."))

        return {
            "value": value,
            "variation": None,
            "source": "BCP",
        }

    except Exception:
        return {"value": None, "variation": None, "source": "BCP"}

def update_external_indicators(conn: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    """
    Actualiza indicadores externos.
    Por ahora solo USD/PYG desde BCP.
    """
    results: dict[str, dict[str, Any]] = {}

    usd = fetch_usd_pyg_bcp()
    results["usd_pyg"] = usd

    if usd["value"] is not None:
        conn.execute(
            """
            UPDATE indicators
            SET value = ?, variation = ?, updated_at = CURRENT_TIMESTAMP
            WHERE key = 'usd_pyg'
            """,
            (usd["value"], usd["variation"]),
        )

    return results


def income_settings(conn: sqlite3.Connection) -> dict[str, Any]:
    row = conn.execute(
        "SELECT label, monthly_income, household_size, updated_at FROM income_settings ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if not row:
        return {
            "label": "Salario mínimo mensual",
            "monthly_income": 2798510.0,
            "household_size": 4.0,
            "updated_at": None,
        }
    return dict(row)



def summary_for_week(conn: sqlite3.Connection, week_date: str | None, location_code: str = "NAT") -> list[dict[str, Any]]:
    if not week_date:
        return []
    prev = previous_week(conn, week_date, location_code)
    rows = conn.execute(
        """
        SELECT
            p.id,
            p.name,
            p.weight,
            p.import_dependency,
            p.fx_sensitivity,
            p.fuel_sensitivity,
            w.price AS current_price,
            pw.price AS previous_price
        FROM weekly_prices w
        JOIN products p ON p.id = w.product_id
        JOIN locations l ON l.id = w.location_id
        LEFT JOIN weekly_prices pw
               ON pw.product_id = w.product_id
              AND pw.location_id = w.location_id
              AND pw.week_date = ?
        WHERE w.week_date = ?
          AND l.code = ?
        ORDER BY p.id
        """,
        (prev, week_date, location_code),
    ).fetchall()

    out: list[dict[str, Any]] = []
    for r in rows:
        cp = float(r["current_price"]) if r["current_price"] is not None else None
        pp = float(r["previous_price"]) if r["previous_price"] is not None else None
        variation = calculate_variation(cp, pp)
        out.append(
            {
                "product_id": int(r["id"]),
                "name": r["name"],
                "weight": float(r["weight"] or 1),
                "import_dependency": float(r["import_dependency"] or 0),
                "fx_sensitivity": float(r["fx_sensitivity"] or 0),
                "fuel_sensitivity": float(r["fuel_sensitivity"] or 0),
                "current_price": cp,
                "previous_price": pp,
                "variation": variation,
            }
        )
    return out



def ranking_for_week(conn: sqlite3.Connection, week_date: str | None, location_code: str = "NAT") -> list[dict[str, Any]]:
    data = summary_for_week(conn, week_date, location_code)
    ranked = [x for x in data if x["variation"] is not None]
    ranked.sort(key=lambda x: x["variation"], reverse=True)
    return ranked


# =========================================================
# MÓDULO IPPS / ALERTAS / SOCIAL / TERRITORIAL
# =========================================================
def inflation_pressure(conn: sqlite3.Connection, location_code: str = "NAT") -> dict[str, Any]:
    week = latest_week(conn, location_code)
    items = summary_for_week(conn, week, location_code)
    if not items:
        return {
            "week": week,
            "location": location_code,
            "ipps": None,
            "level": "Sin datos",
            "diffusion_rate": None,
            "avg_increase": None,
            "weighted_pressure": None,
            "products_up": 0,
            "products_down": 0,
            "products_flat": 0,
            "message": "No hay datos suficientes para calcular IPPS.",
        }

    items_with_var = [x for x in items if x["variation"] is not None]
    if not items_with_var:
        return {
            "week": week,
            "location": location_code,
            "ipps": None,
            "level": "Sin base previa",
            "diffusion_rate": None,
            "avg_increase": None,
            "weighted_pressure": None,
            "products_up": 0,
            "products_down": 0,
            "products_flat": len(items),
            "message": "Existe una sola semana o no hay semana previa comparable.",
        }

    up = [x for x in items_with_var if x["variation"] > 0]
    down = [x for x in items_with_var if x["variation"] < 0]
    flat = [x for x in items_with_var if x["variation"] == 0]
    total = len(items_with_var)

    diffusion_rate = (len(up) / total) * 100 if total else None
    avg_increase = sum(x["variation"] for x in up) / len(up) if up else 0.0

    total_weight = sum(float(x["weight"] or 1) for x in items_with_var) or 1.0
    weighted_pressure = sum(
        max(float(x["variation"] or 0), 0) * float(x["weight"] or 1) for x in items_with_var
    ) / total_weight

    # IPPS simple y potente para gestión: difusión + intensidad + peso
    ipps = (0.45 * (diffusion_rate or 0)) + (0.35 * avg_increase) + (0.20 * weighted_pressure)

    if ipps < 8:
        level = "Normal"
        color = "green"
    elif ipps < 15:
        level = "Presión"
        color = "yellow"
    else:
        level = "Alerta"
        color = "red"

    message = {
        "Normal": "La presión semanal luce contenida.",
        "Presión": "Se observa una difusión relevante de aumentos que merece seguimiento.",
        "Alerta": "La presión de precios es alta y requiere monitoreo prioritario.",
    }[level]

    return {
        "week": week,
        "location": location_code,
        "ipps": round(ipps, 2),
        "level": level,
        "color": color,
        "diffusion_rate": round(diffusion_rate, 2) if diffusion_rate is not None else None,
        "avg_increase": round(avg_increase, 2) if avg_increase is not None else None,
        "weighted_pressure": round(weighted_pressure, 2) if weighted_pressure is not None else None,
        "products_up": len(up),
        "products_down": len(down),
        "products_flat": len(flat),
        "message": message,
    }



def top_explanatory_pressures(conn: sqlite3.Connection) -> dict[str, Any]:
    inds = indicators_map(conn)
    usd_var = to_float(inds.get("usd_pyg", {}).get("variation")) or 0.0
    diesel_var = to_float(inds.get("diesel", {}).get("variation")) or 0.0
    gasolina_var = to_float(inds.get("gasolina", {}).get("variation")) or 0.0
    brent_var = to_float(inds.get("brent", {}).get("variation")) or 0.0
    trigo_var = to_float(inds.get("trigo", {}).get("variation")) or 0.0
    maiz_var = to_float(inds.get("maiz", {}).get("variation")) or 0.0

    fx_pressure = abs(usd_var)
    fuel_pressure = (abs(diesel_var) + abs(gasolina_var) + abs(brent_var)) / 3 if any([diesel_var, gasolina_var, brent_var]) else 0.0
    grain_pressure = (abs(trigo_var) + abs(maiz_var)) / 2 if any([trigo_var, maiz_var]) else 0.0

    return {
        "fx_pressure": round(fx_pressure, 2),
        "fuel_pressure": round(fuel_pressure, 2),
        "grain_pressure": round(grain_pressure, 2),
        "usd_variation": round(usd_var, 2),
        "diesel_variation": round(diesel_var, 2),
        "gasolina_variation": round(gasolina_var, 2),
        "brent_variation": round(brent_var, 2),
        "trigo_variation": round(trigo_var, 2),
        "maiz_variation": round(maiz_var, 2),
    }



def generate_alerts(conn: sqlite3.Connection, location_code: str = "NAT") -> list[dict[str, Any]]:
    week = latest_week(conn, location_code)
    ranking = ranking_for_week(conn, week, location_code)
    ipps = inflation_pressure(conn, location_code)
    external = top_explanatory_pressures(conn)
    alerts: list[dict[str, Any]] = []

    if ipps.get("level") == "Alerta":
        alerts.append(
            {
                "type": "INFLATION_PRESSURE",
                "severity": "ALTA",
                "title": "Alerta por presión semanal de precios",
                "detail": f"IPPS = {ipps.get('ipps')} en {location_code}. {ipps.get('message')}",
                "week": week,
            }
        )
    elif ipps.get("level") == "Presión":
        alerts.append(
            {
                "type": "INFLATION_PRESSURE",
                "severity": "MEDIA",
                "title": "Presión relevante en precios",
                "detail": f"IPPS = {ipps.get('ipps')} en {location_code}.",
                "week": week,
            }
        )

    for row in ranking[:5]:
        if (row.get("variation") or 0) >= 10:
            risk = "ALTA" if row["import_dependency"] >= 0.5 or row["fuel_sensitivity"] >= 0.5 else "MEDIA"
            alerts.append(
                {
                    "type": "PRODUCT_SPIKE",
                    "severity": risk,
                    "title": f"Aumento abrupto en {row['name']}",
                    "detail": f"Variación semanal de {pct(row['variation'])}. Precio actual: {money(row['current_price'])}.",
                    "week": week,
                }
            )

    if external["fx_pressure"] >= 2.5:
        alerts.append(
            {
                "type": "FX_RISK",
                "severity": "MEDIA",
                "title": "Presión cambiaria relevante",
                "detail": f"Variación USD/PYG: {pct(external['usd_variation'])}. Riesgo de traslado a importados.",
                "week": week,
            }
        )

    if external["fuel_pressure"] >= 3:
        alerts.append(
            {
                "type": "FUEL_RISK",
                "severity": "MEDIA",
                "title": "Presión en combustibles y logística",
                "detail": "El bloque combustible muestra variaciones relevantes con posible impacto transversal.",
                "week": week,
            }
        )

    if not alerts:
        alerts.append(
            {
                "type": "STATUS",
                "severity": "BAJA",
                "title": "Sin alertas relevantes",
                "detail": "No se detectan señales críticas con la información disponible.",
                "week": week,
            }
        )

    return alerts



def social_pressure(conn: sqlite3.Connection, location_code: str = "NAT") -> dict[str, Any]:
    week = latest_week(conn, location_code)
    items = summary_for_week(conn, week, location_code)
    income = income_settings(conn)

    if not items:
        return {
            "week": week,
            "location": location_code,
            "basket_cost": None,
            "monthly_income": income["monthly_income"],
            "weekly_income": round(float(income["monthly_income"]) / 4.0, 2),
            "basket_share_of_weekly_income": None,
            "social_pressure_level": "Sin datos",
            "message": "No hay datos suficientes para calcular presión social.",
        }

    # aproximación ejecutiva: costo agregado semanal de canasta observada
    basket_cost = sum(float(x["current_price"] or 0) for x in items)
    monthly_income = float(income["monthly_income"])
    weekly_income = monthly_income / 4.0 if monthly_income else 0.0
    share = (basket_cost / weekly_income) * 100 if weekly_income else None

    if share is None:
        level = "Sin datos"
    elif share < 35:
        level = "Baja"
    elif share < 50:
        level = "Media"
    else:
        level = "Alta"

    message = {
        "Baja": "La canasta observada luce manejable respecto al ingreso de referencia.",
        "Media": "La canasta absorbe una fracción importante del ingreso semanal.",
        "Alta": "La canasta observada ejerce una presión alta sobre el ingreso semanal.",
        "Sin datos": "Sin información suficiente.",
    }[level]

    return {
        "week": week,
        "location": location_code,
        "basket_cost": round(basket_cost, 2),
        "monthly_income": round(monthly_income, 2),
        "weekly_income": round(weekly_income, 2),
        "basket_share_of_weekly_income": round(share, 2) if share is not None else None,
        "social_pressure_level": level,
        "message": message,
    }



def territorial_prices(conn: sqlite3.Connection, product_name: str | None = None, week_date: str | None = None) -> dict[str, Any]:
    week = week_date or latest_week(conn, "NAT")
    if not week:
        return {"week": week, "items": []}

    params: list[Any] = [week]
    product_filter = ""
    if product_name:
        product_filter = " AND p.normalized_name = ? "
        params.append(normalize_name(product_name))

    rows = conn.execute(
        f"""
        SELECT l.code AS location_code, l.name AS location_name, p.name AS product_name, w.price
        FROM weekly_prices w
        JOIN products p ON p.id = w.product_id
        JOIN locations l ON l.id = w.location_id
        WHERE w.week_date = ?
        {product_filter}
        ORDER BY p.name, l.id
        """,
        params,
    ).fetchall()

    items = [
        {
            "location_code": r["location_code"],
            "location_name": r["location_name"],
            "product_name": r["product_name"],
            "price": float(r["price"]),
        }
        for r in rows
    ]
    return {"week": week, "items": items}


# =========================================================
# IMPORTACIÓN EXCEL SEGURA
# =========================================================
class ImportResult(BaseModel):
    ok: bool
    filename: str
    rows_detected: int = 0
    rows_inserted: int = 0
    weeks_detected: int = 0
    products_detected: int = 0
    location_code: str = "NAT"
    message: str



def ensure_product(conn: sqlite3.Connection, raw_name: str, col_index: int) -> int:
    nname = normalize_name(raw_name)
    if not nname:
        raise ValueError(f"Encabezado vacío en columna {col_index}")

    row = conn.execute("SELECT id FROM products WHERE normalized_name = ?", (nname,)).fetchone()
    if row:
        return int(row["id"])

    count = conn.execute("SELECT COUNT(*) AS n FROM products").fetchone()["n"]
    code = f"P{int(count) + 1:02d}"
    cur = conn.execute(
        """
        INSERT INTO products (code, name, normalized_name, unit, weight, import_dependency, fx_sensitivity, fuel_sensitivity, active)
        VALUES (?, ?, ?, 'unidad', 1, 0.20, 0.20, 0.20, 1)
        """,
        (code, raw_name.strip(), nname),
    )
    return int(cur.lastrowid)



def parse_canasta_excel(conn: sqlite3.Connection, excel_path: Path, source: str, location_code: str = "NAT") -> ImportResult:

    wb = load_workbook(excel_path, data_only=True)

    # decidir hoja automáticamente
    if "Carga_Semanal" in wb.sheetnames:
        sh = wb["Carga_Semanal"]
        mode = "long"
    elif "Canasta_25" in wb.sheetnames:
        sh = wb["Canasta_25"]
        mode = "pivot"
    else:
        sh = wb[wb.sheetnames[0]]
        mode = "pivot"

    location_id = get_location_id(conn, location_code)

    staged_rows = []
    weeks = set()
    detected_products = set()

    # ======================================================
    # FORMATO 1: Carga_Semanal (Google Sheets)
    # columnas: fecha_semana | producto | precio
    # ======================================================
    if mode == "long":

        headers = {}

        for col in range(1, sh.max_column + 1):
            val = sh.cell(row=1, column=col).value
            if val:
                headers[str(val).strip().lower()] = col

        if "fecha_semana" not in headers or "producto" not in headers or "precio" not in headers:
            raise ValueError("La hoja Carga_Semanal debe tener columnas: fecha_semana, producto, precio")

        col_fecha = headers["fecha_semana"]
        col_producto = headers["producto"]
        col_precio = headers["precio"]

        for r in range(2, sh.max_row + 1):

            week_date = iso_date(sh.cell(row=r, column=col_fecha).value)
            product = sh.cell(row=r, column=col_producto).value
            price = to_float(sh.cell(row=r, column=col_precio).value)

            if not week_date or not product or price is None:
                continue

            product_id = ensure_product(conn, str(product), col_producto)

            staged_rows.append((week_date, product_id, location_id, price, source))

            weeks.add(week_date)
            detected_products.add(product_id)

    # ======================================================
    # FORMATO 2: Canasta_25 (formato pivot del observatorio)
    # ======================================================
    else:

        product_map = {}

        for col in range(2, 27):

            header = sh.cell(row=6, column=col).value

            if header:
                product_id = ensure_product(conn, str(header), col)
                product_map[col] = product_id

        for r in range(8, sh.max_row + 1):

            week_date = iso_date(sh.cell(row=r, column=1).value)

            if not week_date:
                continue

            for col, product_id in product_map.items():

                price = to_float(sh.cell(row=r, column=col).value)

                if price is None:
                    continue

                staged_rows.append((week_date, product_id, location_id, price, source))

                weeks.add(week_date)
                detected_products.add(product_id)

    if not staged_rows:
        raise ValueError("No se detectaron precios válidos para importar.")

    inserted = 0

    for week_date, product_id, loc_id, price, src in staged_rows:

        conn.execute(
            """
            INSERT INTO weekly_prices (week_date, product_id, location_id, price, source)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(week_date, product_id, location_id)
            DO UPDATE SET
                price = excluded.price,
                source = excluded.source,
                imported_at = CURRENT_TIMESTAMP
            """,
            (week_date, product_id, loc_id, price, src),
        )

        inserted += 1

    return ImportResult(
        ok=True,
        filename=excel_path.name,
        rows_detected=len(staged_rows),
        rows_inserted=inserted,
        weeks_detected=len(weeks),
        products_detected=len(detected_products),
        location_code=location_code,
        message="Importación realizada correctamente.",
    )

# =========================================================
# HTML DASHBOARD
# =========================================================
def render_dashboard(conn: sqlite3.Connection, message: str = "") -> str:
    week = latest_week(conn, "NAT")
    summary = summary_for_week(conn, week, "NAT")
    ranking = ranking_for_week(conn, week, "NAT")[:10]
    inds = indicators_map(conn)
    ipps = inflation_pressure(conn, "NAT")
    social = social_pressure(conn, "NAT")
    alerts = generate_alerts(conn, "NAT")[:5]

    last_import = conn.execute(
        "SELECT filename, status, message, created_at FROM imports_log ORDER BY id DESC LIMIT 1"
    ).fetchone()

    rows_html = ""
    for row in summary:
        var_class = "up" if (row["variation"] or 0) > 0 else "down" if (row["variation"] or 0) < 0 else "flat"
        rows_html += f"""
        <tr>
            <td>{row['name']}</td>
            <td>{money(row['current_price'])}</td>
            <td>{money(row['previous_price'])}</td>
            <td class="{var_class}">{pct(row['variation'])}</td>
        </tr>
        """

    ranking_html = ""
    for idx, row in enumerate(ranking, start=1):
        ranking_html += f"""
        <tr>
            <td>{idx}</td>
            <td>{row['name']}</td>
            <td>{pct(row['variation'])}</td>
            <td>{money(row['current_price'])}</td>
        </tr>
        """

    if not rows_html:
        rows_html = '<tr><td colspan="4">No hay datos cargados todavía.</td></tr>'
    if not ranking_html:
        ranking_html = '<tr><td colspan="4">No hay ranking disponible todavía.</td></tr>'

    def ind_card(key: str) -> str:
        obj = inds.get(key, {})
        value = obj.get("value")
        variation = obj.get("variation")
        label = obj.get("label", key)
        return f"""
        <div class="card metric">
            <div class="metric-title">{label}</div>
            <div class="metric-value">{value if value is not None else '—'}</div>
            <div class="metric-sub">Variación: {pct(variation)}</div>
        </div>
        """

    import_info = ""
    if last_import:
        import_info = (
            f"Última importación: {last_import['created_at']} · {last_import['filename'] or 's/archivo'} · "
            f"{last_import['status']} · {last_import['message'] or ''}"
        )

    alerts_html = "".join(
        f"<li><strong>{a['severity']}</strong> — {a['title']}: {a['detail']}</li>" for a in alerts
    ) or "<li>Sin alertas.</li>"

    latest_week_label = week or "Sin datos"
    ipps_level = ipps.get("level", "Sin datos")
    ipps_value = ipps.get("ipps") if ipps.get("ipps") is not None else "—"

    return f"""
    <!doctype html>
    <html lang="es">
    <head>
        <meta charset="utf-8" />
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <title>{APP_TITLE}</title>
        <style>
            :root {{
                --bg: #f4f7fb;
                --card: #ffffff;
                --primary: #113a69;
                --primary-2: #285f9d;
                --text: #1f2937;
                --muted: #6b7280;
                --border: #dbe3ee;
                --up: #0f8a43;
                --down: #b42318;
                --flat: #6b7280;
            }}
            * {{ box-sizing: border-box; }}
            body {{ margin: 0; font-family: Arial, Helvetica, sans-serif; background: var(--bg); color: var(--text); }}
            .wrap {{ max-width: 1280px; margin: 0 auto; padding: 24px; }}
            .hero {{ background: linear-gradient(135deg, var(--primary), var(--primary-2)); color: white; padding: 24px; border-radius: 18px; }}
            .hero h1 {{ margin: 0 0 8px 0; font-size: 28px; }}
            .hero p {{ margin: 0; opacity: .95; }}
            .grid {{ display: grid; gap: 16px; }}
            .grid-2 {{ grid-template-columns: 1.15fr .85fr; }}
            .grid-3 {{ grid-template-columns: repeat(3, 1fr); }}
            .grid-4 {{ grid-template-columns: repeat(4, 1fr); }}
            .grid-6 {{ grid-template-columns: repeat(6, 1fr); }}
            .card {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; padding: 18px; box-shadow: 0 6px 18px rgba(17,58,105,.06); }}
            .card h2 {{ margin: 0 0 14px; font-size: 20px; color: var(--primary); }}
            .toolbar {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
            .btn {{ display: inline-block; border: 0; padding: 10px 14px; border-radius: 10px; background: var(--primary); color: white; text-decoration: none; cursor: pointer; font-weight: 700; }}
            .btn.alt {{ background: #e8eef7; color: var(--primary); }}
            .file {{ padding: 10px; border: 1px solid var(--border); border-radius: 10px; background: white; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ text-align: left; padding: 10px 8px; border-bottom: 1px solid var(--border); font-size: 14px; }}
            th {{ color: var(--primary); background: #f8fbff; }}
            .metric {{ min-height: 128px; }}
            .metric-title {{ font-size: 13px; color: var(--muted); margin-bottom: 8px; }}
            .metric-value {{ font-size: 28px; font-weight: 800; color: var(--primary); }}
            .metric-sub {{ margin-top: 8px; color: var(--muted); font-size: 13px; }}
            .muted {{ color: var(--muted); font-size: 13px; }}
            .up {{ color: var(--up); font-weight: 700; }}
            .down {{ color: var(--down); font-weight: 700; }}
            .flat {{ color: var(--flat); font-weight: 700; }}
            .notice {{ margin: 14px 0 0; padding: 12px 14px; border-radius: 12px; background: #eef6ff; border: 1px solid #cfe0f5; color: var(--primary); }}
            ul.alerts {{ padding-left: 18px; margin: 0; }}
            ul.alerts li {{ margin: 8px 0; }}
            @media (max-width: 1100px) {{
                .grid-6 {{ grid-template-columns: repeat(3, 1fr); }}
                .grid-4 {{ grid-template-columns: repeat(2, 1fr); }}
                .grid-2 {{ grid-template-columns: 1fr; }}
            }}
            @media (max-width: 700px) {{
                .grid-6, .grid-4, .grid-3 {{ grid-template-columns: 1fr; }}
                .wrap {{ padding: 14px; }}
            }}
        </style>
    </head>
    <body>
        <div class="wrap">
            <div class="hero">
                <h1>OMPP Sistema con Reporte Real — Dashboard</h1>
                <p>Observatorio de precios, presión inflacionaria, riesgo y monitoreo social.</p>
            </div>

            {f'<div class="notice">{message}</div>' if message else ''}

            <div class="grid grid-2" style="margin-top:16px;">
                <div class="card">
                    <h2>Herramientas</h2>
                    <div class="toolbar" style="margin-bottom:12px;">
                        <a class="btn alt" href="/ranking">Ver Ranking</a>
                        <a class="btn alt" href="/report/pdf">Descargar reporte PDF</a>
                        <a class="btn alt" href="/backup.csv">Exportar CSV</a>
                    </div>
                    <form action="/import_excel" method="post" enctype="multipart/form-data">
    <div class="toolbar" style="display:flex; gap:12px; align-items:center; flex-wrap:wrap;">
        <input type="file" name="file" accept=".xlsx,.xlsm" required style="padding:8px; background:white; border:1px solid #dbe3ee; border-radius:8px;" />
        <input type="hidden" name="source" value="Excel" />
        <input type="hidden" name="location_code" value="NAT" />
        <button class="btn" type="submit">Importar Excel</button>
    </div>
</form>
                    <p class="muted" style="margin-top:12px;">Última semana disponible: <strong>{latest_week_label}</strong></p>
                    <p class="muted">{import_info}</p>
                </div>

                <div class="card">
                    <h2>Indicadores Internacionales</h2>
                    <div class="grid grid-3">
                        {ind_card('usd_pyg')}
                        {ind_card('brent')}
                        {ind_card('diesel')}
                        {ind_card('gasolina')}
                        {ind_card('trigo')}
                        {ind_card('maiz')}
                    </div>
                </div>
            </div>

            <div class="grid grid-4" style="margin-top:16px;">
                <div class="card metric">
                    <div class="metric-title">IPPS</div>
                    <div class="metric-value">{ipps_value}</div>
                    <div class="metric-sub">Nivel: {ipps_level}</div>
                </div>
                <div class="card metric">
                    <div class="metric-title">Productos con suba</div>
                    <div class="metric-value">{ipps.get('products_up', 0)}</div>
                    <div class="metric-sub">Difusión: {pct(ipps.get('diffusion_rate')) if ipps.get('diffusion_rate') is not None else '—'}</div>
                </div>
                <div class="card metric">
                    <div class="metric-title">Costo canasta observada</div>
                    <div class="metric-value">{money(social.get('basket_cost'))}</div>
                    <div class="metric-sub">Ingreso semanal ref.: {money(social.get('weekly_income'))}</div>
                </div>
                <div class="card metric">
                    <div class="metric-title">Presión social</div>
                    <div class="metric-value">{social.get('social_pressure_level', '—')}</div>
                    <div class="metric-sub">Participación en ingreso: {pct(social.get('basket_share_of_weekly_income')) if social.get('basket_share_of_weekly_income') is not None else '—'}</div>
                </div>
            </div>

            <div class="grid grid-2" style="margin-top:16px;">
                <div class="card">
                    <h2>Precios semanales</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>Producto</th>
                                <th>Semana actual</th>
                                <th>Semana previa</th>
                                <th>Variación</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>

                <div class="card">
                    <h2>Ranking de variaciones</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>#</th>
                                <th>Producto</th>
                                <th>Variación</th>
                                <th>Precio actual</th>
                            </tr>
                        </thead>
                        <tbody>
                            {ranking_html}
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="grid grid-2" style="margin-top:16px;">
                <div class="card">
                    <h2>Alertas tempranas</h2>
                    <ul class="alerts">{alerts_html}</ul>
                </div>
                <div class="card">
                    <h2>API del Observatorio</h2>
                    <table>
                        <tbody>
                            <tr><td>/api/summary</td><td>Resumen semanal</td></tr>
                            <tr><td>/api/ranking</td><td>Ranking semanal</td></tr>
                            <tr><td>/api/inflation_pressure</td><td>IPPS</td></tr>
                            <tr><td>/api/alerts</td><td>Alertas</td></tr>
                            <tr><td>/api/territorial_prices</td><td>Precios por territorio</td></tr>
                            <tr><td>/api/social_pressure</td><td>Presión sobre ingreso</td></tr>
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    """


# =========================================================
# ENDPOINTS HTML
# =========================================================
@app.get("/", response_class=HTMLResponse)
def home() -> HTMLResponse:
    with db() as conn:
        try:
            update_external_indicators(conn)
        except Exception:
            pass
        return HTMLResponse(render_dashboard(conn))


@app.post("/import_excel", response_class=HTMLResponse)
async def import_excel(
    file: UploadFile = File(...),
    source: str = Form(DEFAULT_SOURCE),
    location_code: str = Form("NAT"),
) -> HTMLResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No se recibió archivo.")

    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Formato no soportado. Usa .xlsx o .xlsm")

    safe_name = f"{dt.datetime.now():%Y%m%d_%H%M%S}_{Path(file.filename).name}"
    save_path = UPLOAD_DIR / safe_name
    content = await file.read()
    save_path.write_bytes(content)

    with db() as conn:
        try:
            result = parse_canasta_excel(conn, save_path, source, location_code=location_code)
            conn.execute(
                """
                INSERT INTO imports_log (filename, status, message, rows_detected, rows_inserted)
                VALUES (?, 'OK', ?, ?, ?)
                """,
                (result.filename, result.message, result.rows_detected, result.rows_inserted),
            )
            message = (
                f"Importación correcta: {result.rows_inserted} registros, "
                f"{result.weeks_detected} semanas, {result.products_detected} productos, ubicación {result.location_code}."
            )
            return HTMLResponse(render_dashboard(conn, message=message))
        except Exception as e:
            conn.execute(
                """
                INSERT INTO imports_log (filename, status, message, rows_detected, rows_inserted)
                VALUES (?, 'ERROR', ?, 0, 0)
                """,
                (safe_name, str(e)),
            )
            return HTMLResponse(render_dashboard(conn, message=f"Error al importar: {e}"), status_code=400)


@app.get("/ranking", response_class=HTMLResponse)
def ranking_view() -> HTMLResponse:
    with db() as conn:
        week = latest_week(conn, "NAT")
        ranking = ranking_for_week(conn, week, "NAT")
        rows = "".join(
            f"<tr><td>{i}</td><td>{r['name']}</td><td>{pct(r['variation'])}</td><td>{money(r['current_price'])}</td><td>{money(r['previous_price'])}</td></tr>"
            for i, r in enumerate(ranking, start=1)
        )
        if not rows:
            rows = '<tr><td colspan="5">No hay ranking disponible todavía.</td></tr>'
        html = f"""
        <!doctype html>
        <html lang="es"><head><meta charset="utf-8"><title>Ranking</title>
        <style>
            body{{font-family:Arial;background:#f4f7fb;padding:24px;color:#1f2937}}
            .card{{max-width:1100px;margin:auto;background:#fff;border:1px solid #dbe3ee;border-radius:16px;padding:18px}}
            a{{text-decoration:none;color:#113a69;font-weight:bold}}
            table{{width:100%;border-collapse:collapse;margin-top:12px}}
            th,td{{padding:10px;border-bottom:1px solid #dbe3ee;text-align:left}}
            th{{background:#f8fbff;color:#113a69}}
        </style></head>
        <body><div class="card">
            <p><a href="/">← Volver al dashboard</a></p>
            <h1>Ranking de variaciones</h1>
            <p>Semana analizada: <strong>{week or 'Sin datos'}</strong></p>
            <table>
                <thead><tr><th>#</th><th>Producto</th><th>Variación</th><th>Precio actual</th><th>Precio previo</th></tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </div></body></html>
        """
        return HTMLResponse(html)


# =========================================================
# API
# =========================================================
@app.get("/api/summary")
def api_summary(location_code: str = "NAT") -> JSONResponse:
    with db() as conn:
        week = latest_week(conn, location_code)
        return JSONResponse({"week": week, "location": location_code, "items": summary_for_week(conn, week, location_code)})

@app.get("/api/ranking")
def api_ranking(location_code: str = "NAT") -> JSONResponse:
    with db() as conn:
        week = latest_week(conn, location_code)
        return JSONResponse({"week": week, "location": location_code, "items": ranking_for_week(conn, week, location_code)})


@app.get("/api/inflation_pressure")
def api_inflation_pressure(location_code: str = "NAT") -> JSONResponse:
    with db() as conn:
        return JSONResponse(inflation_pressure(conn, location_code))


@app.get("/api/alerts")
def api_alerts(location_code: str = "NAT") -> JSONResponse:
    with db() as conn:
        return JSONResponse({"location": location_code, "items": generate_alerts(conn, location_code)})


@app.get("/api/territorial_prices")
def api_territorial_prices(product: str | None = None, week_date: str | None = None) -> JSONResponse:
    with db() as conn:
        return JSONResponse(territorial_prices(conn, product_name=product, week_date=week_date))


@app.get("/api/social_pressure")
def api_social_pressure(location_code: str = "NAT") -> JSONResponse:
    with db() as conn:
        return JSONResponse(social_pressure(conn, location_code))

@app.get("/api/update_external_indicators")
def api_update_external_indicators() -> JSONResponse:
    with db() as conn:
        results = update_external_indicators(conn)
        return JSONResponse({"ok": True, "results": results})

@app.post("/api/indicators")
async def api_update_indicators(request: Request) -> JSONResponse:
    payload = await request.json()
    allowed = {"usd_pyg", "brent", "diesel", "gasolina", "trigo", "maiz"}
    updated = []
    with db() as conn:
        for key, value in payload.items():
            if key not in allowed or not isinstance(value, dict):
                continue
            conn.execute(
                """
                UPDATE indicators
                SET value = ?, variation = ?, updated_at = CURRENT_TIMESTAMP
                WHERE key = ?
                """,
                (to_float(value.get("value")), to_float(value.get("variation")), key),
            )
            updated.append(key)
    return JSONResponse({"ok": True, "updated": updated})


@app.post("/api/income_settings")
async def api_income_settings(request: Request) -> JSONResponse:
    payload = await request.json()
    monthly_income = to_float(payload.get("monthly_income"))
    household_size = to_float(payload.get("household_size")) or 4
    label = str(payload.get("label") or "Salario mínimo mensual")
    if monthly_income is None or monthly_income <= 0:
        raise HTTPException(status_code=400, detail="monthly_income debe ser un número mayor a 0")

    with db() as conn:
        conn.execute(
            "INSERT INTO income_settings (label, monthly_income, household_size) VALUES (?, ?, ?)",
            (label, monthly_income, household_size),
        )
    return JSONResponse({"ok": True, "message": "Parámetros sociales actualizados."})


@app.post("/api/territorial_price")
async def api_add_territorial_price(request: Request) -> JSONResponse:
    payload = await request.json()
    week_date = iso_date(payload.get("week_date"))
    product_name = str(payload.get("product_name") or "").strip()
    location_code = str(payload.get("location_code") or "NAT").strip().upper()
    price = to_float(payload.get("price"))
    source = str(payload.get("source") or "manual")

    if not week_date or not product_name or price is None:
        raise HTTPException(status_code=400, detail="week_date, product_name y price son obligatorios")

    with db() as conn:
        location_id = get_location_id(conn, location_code)
        product_id = ensure_product(conn, product_name, 0)
        conn.execute(
            """
            INSERT INTO weekly_prices (week_date, product_id, location_id, price, source)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(week_date, product_id, location_id)
            DO UPDATE SET price=excluded.price, source=excluded.source, imported_at=CURRENT_TIMESTAMP
            """,
            (week_date, product_id, location_id, price, source),
        )
    return JSONResponse({"ok": True, "message": "Precio territorial guardado."})


@app.get("/backup.csv")
def backup_csv() -> Response:
    with db() as conn:
        rows = conn.execute(
            """
            SELECT w.week_date, l.code AS location_code, l.name AS location_name,
                   p.code, p.name, p.unit, w.price, w.source, w.imported_at
            FROM weekly_prices w
            JOIN products p ON p.id = w.product_id
            JOIN locations l ON l.id = w.location_id
            ORDER BY w.week_date DESC, l.id ASC, p.id ASC
            """
        ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["week_date", "location_code", "location_name", "code", "product", "unit", "price", "source", "imported_at"])
    for r in rows:
        writer.writerow([
            r["week_date"], r["location_code"], r["location_name"], r["code"],
            r["name"], r["unit"], r["price"], r["source"], r["imported_at"]
        ])

    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="ompp_backup.csv"'},
    )


# =========================================================
# PDF
# =========================================================
@app.get("/report/pdf")
def report_pdf() -> FileResponse:
    with db() as conn:
        week = latest_week(conn, "NAT")
        summary = summary_for_week(conn, week, "NAT")
        ranking = ranking_for_week(conn, week, "NAT")[:10]
        ipps = inflation_pressure(conn, "NAT")
        social = social_pressure(conn, "NAT")
        alerts = generate_alerts(conn, "NAT")[:5]

    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_pdf.close()

    doc = SimpleDocTemplate(
        temp_pdf.name,
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleOMPP",
        parent=styles["Title"],
        textColor=colors.HexColor("#113A69"),
        fontSize=18,
        leading=22,
        spaceAfter=10,
    )
    normal = styles["BodyText"]

    elements = []
    elements.append(Paragraph("OMPP — Reporte Semanal de Canasta Básica", title_style))
    elements.append(Paragraph(f"Semana de referencia: <b>{week or 'Sin datos'}</b>", normal))
    elements.append(Paragraph(f"Fecha de emisión: <b>{dt.datetime.now():%d/%m/%Y %H:%M}</b>", normal))
    elements.append(Paragraph(f"IPPS: <b>{ipps.get('ipps') if ipps.get('ipps') is not None else '—'}</b> · Nivel: <b>{ipps.get('level', 'Sin datos')}</b>", normal))
    elements.append(Paragraph(f"Presión social: <b>{social.get('social_pressure_level', 'Sin datos')}</b> · Participación de canasta en ingreso semanal: <b>{pct(social.get('basket_share_of_weekly_income')) if social.get('basket_share_of_weekly_income') is not None else '—'}</b>", normal))
    elements.append(Spacer(1, 0.4 * cm))

    table_data = [["Producto", "Precio actual", "Precio previo", "Variación"]]
    if summary:
        for row in summary:
            table_data.append([row["name"], money(row["current_price"]), money(row["previous_price"]), pct(row["variation"])])
    else:
        table_data.append(["No hay datos", "—", "—", "—"])

    tbl = Table(table_data, repeatRows=1, colWidths=[6.5 * cm, 3.2 * cm, 3.2 * cm, 2.6 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#113A69")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7E2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFE")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 0.5 * cm))

    rank_data = [["#", "Producto", "Variación"]]
    if ranking:
        for idx, row in enumerate(ranking, start=1):
            rank_data.append([str(idx), row["name"], pct(row["variation"])])
    else:
        rank_data.append(["—", "Sin ranking", "—"])

    rank_tbl = Table(rank_data, repeatRows=1, colWidths=[1.2 * cm, 10.3 * cm, 3.5 * cm])
    rank_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D5F8B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7E2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFE")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(Paragraph("Top 10 variaciones semanales", styles["Heading2"]))
    elements.append(rank_tbl)
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph("Alertas tempranas", styles["Heading2"]))
    for a in alerts:
        elements.append(Paragraph(f"• <b>{a['severity']}</b> — {a['title']}: {a['detail']}", normal))

    doc.build(elements)
    filename = f"reporte_ompp_{week or 'sin_datos'}.pdf"
    return FileResponse(temp_pdf.name, media_type="application/pdf", filename=filename)


# =========================================================
# HEALTH / JOB
# =========================================================
@app.get("/jobs/import_daily")
def import_daily(token: str = "") -> JSONResponse:
    if not JOB_TOKEN or token != JOB_TOKEN:
        raise HTTPException(status_code=403, detail="Token inválido")
    return JSONResponse({"ok": True, "message": "Job activo. Configura aquí tu importación automática segura."})


@app.get("/health")
def health() -> JSONResponse:
    with db() as conn:
        nat_week = latest_week(conn, "NAT")
        count = conn.execute("SELECT COUNT(*) AS n FROM weekly_prices").fetchone()["n"]
        return JSONResponse({
            "ok": True,
            "latest_week": nat_week,
            "rows": count,
            "ipps": inflation_pressure(conn, "NAT"),
            "social": social_pressure(conn, "NAT"),
        })


# =========================================================
# EJECUCIÓN LOCAL
# =========================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")), reload=True)

