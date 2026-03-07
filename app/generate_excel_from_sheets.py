import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import os

# ID DE TU GOOGLE SHEET
SHEET_ID = 165RVQOaLd4UTug40v8dUGuc3W_ZXyCl1hYw46paGHTw/edit?usp=sharing

# URL para descargar
URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv"

print("Descargando datos de Google Sheets...")

df = pd.read_csv(URL)

print("Datos descargados:", len(df), "filas")

# Crear tabla pivot
pivot = df.pivot_table(
    index="fecha_semana",
    columns="producto",
    values="precio",
    aggfunc="mean"
)

pivot = pivot.reset_index()

print("Tabla pivot creada")

# Crear Excel formato Canasta_25
wb = Workbook()
ws = wb.active
ws.title = "Canasta_25"

productos = list(pivot.columns[1:])

# Encabezados fila 6
for i, prod in enumerate(productos):
    ws.cell(row=6, column=i+2).value = prod

# Datos desde fila 8
for r, row in pivot.iterrows():

    ws.cell(row=8+r, column=1).value = row["fecha_semana"]

    for c, prod in enumerate(productos):
        ws.cell(row=8+r, column=2+c).value = row[prod]

# Crear carpeta si no existe
os.makedirs("data/auto_import", exist_ok=True)

file_name = f"data/auto_import/canasta_{datetime.now().date()}.xlsx"

wb.save(file_name)

print("Excel generado:", file_name)
